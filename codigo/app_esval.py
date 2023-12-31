#Librerias sistema
import os
import time
import shutil
import pandas as pd

#Librerias datos
import fitz
import glob
from openpyxl import load_workbook

#Librerias Scrapping
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert

class Scraper_Esval():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None
        
    def diccionario(self):
        dic_mes = {'01':'Enero','02':'Febrero',
            '03':'Marzo','04':'Abril',
            '05':'Mayo','06':'Junio',
            '07':'Julio','08':'Agosto',
            '09':'Septiembre','10':'Octubre',
            '11':'Noviembre','12':'Diciembre',
            }
        return dic_mes
        
    def get_downloads_folder(self):
        # Obtener la ruta del directorio de inicio del usuario actual
        home_dir = os.path.expanduser("~")

        # Combinar el directorio de inicio con la carpeta de descargas para obtener la ruta completa
        downloads_folder = os.path.join(home_dir, "Downloads")

        return downloads_folder
    
    def login(self):
    
        #driver_exe = 'C:\\roda\\esval-portal\\domain\\chromedriver.exe'
        driver_exe = './chromedriver.exe'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        #Version 4.9.1 de selenium
        options = webdriver.ChromeOptions()    
        options.add_argument("--disable-notifications")
        self.driver = webdriver.Chrome(self.driver_path,options=options)

        self.driver.get(self.url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(40)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        #Botones ID y XPATH que utilizaremos para logear
        selector_username_input = 'rut'
        selector_password_input = 'contrasena'
        selector_ingreso_button = 'btn-primary'

        #Encontrar la linea de usuario y setear usuario
        intentos = 0
        set_usuario = True
        while (set_usuario):
                try:
                    print('Try en la funcion usuario...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_username = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.ID, selector_username_input)))
                    element_username.clear()
                    element_username.click()
                    element_username.send_keys(email)
                    set_usuario = False
                except:    
                    print('Exception en la funcion usuario...')
                    print('----------------------------------------------------------------------')
                    set_usuario = intentos <= 3  
            
        #Encontrar la linea de clave y setear clave
        intentos = 0
        set_clave = True
        while (set_clave):
                try:
                    print('Try en la funcion clave ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_password= WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.ID, selector_password_input)))
                    element_password.clear()
                    element_password.click()
                    element_password.send_keys(password)
                    time.sleep(2)
                    set_clave = False
                except:    
                    print('Exception en la funcion clave ...')
                    print('----------------------------------------------------------------------')
                    set_clave = intentos <= 3
        
        #Hacemos click en el boton de ingreso
        intentos = 0
        ingreso = True
        while (ingreso):
                try:
                    print('Try en el boton de ingreso ...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    boton_ingreso = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, selector_ingreso_button)))
                    boton_ingreso.click()
                    ingreso = False
                except:    
                    print('Exception en el boton de ingreso...')
                    print('----------------------------------------------------------------------')
                    ingreso = intentos <= 3        
        
        print('Ya logramos ingresar, ahora vamos a buscar las factruras')
        time.sleep(7)
        
    def scrapping_esval(self):
        print('Entrando en la funcion Scrapping...')
        print('----------------------------------------------------------------------')
        
        folder_path_config = './config/'
        
        # Especifica la ruta de tu archivo Excel
        excel_file = folder_path_config + "clientes.xlsx"

        # Especifica el nombre de la hoja en la que se encuentran los datos
        hoja_excel = "Hoja1"

        # Carga los datos de Excel en un DataFrame
        df = pd.read_excel(excel_file, sheet_name=hoja_excel)
        print('Dataframe ', df)
        print('----------------------------------------------------------------------')

        #Buscamos el menu de las boletas para hacer click
        intentos = 0
        menu_boletas= True
        while (menu_boletas):
            try:
                print('Try localizando el menu de las boletas ...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                boletas = WebDriverWait(self.driver, 30).until(
                EC.element_to_be_clickable((By.XPATH,'/html/body/app-root/app-contenedor-usuarios/div[3]/div/div/div[1]/app-lateral-menu/div/ul/li[2]/a')))
                boletas.click()
                menu_boletas = False
            except:    
                print('Exception en el menu de las boletas ...')
                print('----------------------------------------------------------------------')
                menu_boletas = intentos <= 3    
                    
        #Dentro del menu de boletas buscamos el boton de mis cobros
        intentos = 0
        mis_cobros = True
        while (mis_cobros):
            try:
                print('Try localizando mis cobros dentro del menu boletas ...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                cobros = WebDriverWait(self.driver, 30).until(
                EC.element_to_be_clickable((By.XPATH,'/html/body/app-root/app-contenedor-usuarios/div[3]/div/div/div[1]/app-lateral-menu/div/ul/li[2]/div/ul/li[1]/ul/li[1]/a')))
                cobros.click()
                mis_cobros = False
            except:    
                print('Exception localizando mis cobros dentro del menu boletas ...')
                print('----------------------------------------------------------------------')
                mis_cobros = intentos <= 3    

        #Hacemos doble click sobre expansion para obtener el largo completo de las boletas
        largo_descarga = 13
        
        i = 1            
        while i < largo_descarga+1:

            #Ahora que tenemos el largo completo volvemos a subir para identificar el primera fila y comenzar descarga
            if i == 6 or i==10:
                self.driver.execute_script("window.scrollBy(0, 200)")
                botone_expandir= self.driver.find_element(By.XPATH, '//*[@id="tablaFacturacion"]/div/button/span')
                time.sleep(10)
                botone_expandir.click()
                        
            #Obtenemos la fecha y dividimos dicho valor para tener mes y año            
            reintentos = 0            
            fecha_oficial = True
            while (fecha_oficial) and reintentos <= 5:
                try:
                    print('Try localizando fecha en la tabla ...', reintentos)
                    print('----------------------------------------------------------------------')
                    reintentos += 1
                    fecha_completa = self.driver.find_element(By.XPATH,f'//*[@id="tablaFacturacion"]/div/table/tbody/tr[{i}]/td[1]')
                    fecha_texto = fecha_completa.text
                    partes = fecha_texto.split("/")
                    mes = str(partes[1])
                    año = str(partes[2])
                    print('Hemos localizando fecha en la tabla ...', reintentos)
                    print('----------------------------------------------------------------------')
                    fecha_oficial = False
                except:    
                    print('Exception localizando fecha en la tabla...')
                    print('----------------------------------------------------------------------')
                    reintentos += 1

            #Obtenemos el boton de descarga que haremos click           
            reintentos = 0            
            descarga_oficial = True
            while (descarga_oficial) and reintentos <= 5:
                try:
                    print('Try localizando boton descarga ...', reintentos)
                    print('----------------------------------------------------------------------')
                    reintentos += 1
                    descarga = self.driver.find_element(By.XPATH,f'//*[@id="tablaFacturacion"]/div/table/tbody/tr[{i}]/td[5]/button[3]')
                    descarga.click()
                    time.sleep(7)
                    descarga_oficial = False
                except:    
                    print('Exception en el boton descarga...')
                    print('----------------------------------------------------------------------')
                    reintentos += 1

            # Esperar hasta que el elemento esté presente en la página
            self.driver.implicitly_wait(35)
            descargado = True
            while descargado:                
                # Obtener la URL de la ventana emergente
                downloads_folder = self.get_downloads_folder()

                #Definimos ruta de destino del nuevo archivo
                folder_path = './input/'

                #Buscamos el archivo que se haya descargado que comience con el nombre boleta para poder moverlo a la carpeta input
                for filename in os.listdir(downloads_folder):
                    if filename.startswith("Boleta"):
                        nombre_archivo = filename
                print(nombre_archivo)
                #Split de nombre_archivopara extraer Nro. Boleta
                nombre, extension = nombre_archivo.split(".")

                # Utilizando isdigit() para extraer solo los números
                bill_number = "".join(filter(str.isdigit, nombre))                
                
                #Obtengo resultado del diccionario
                resultado_diccionario = self.diccionario()

                # Cruce datos faltantes para ontener
                for index, row in df.iterrows():
                    
                    df_nro_cliente = df.loc[index, 'nro_cliente']
                    df_sucursal = df.loc[index, 'sucursal']
                    print('Nro. Cliente: ', df_nro_cliente, 'Sucursal: ', df_sucursal)
                    print('--------------------------------------------------------------------------')

                #Defino nuevo nombre para archivo que moveremos al Input
                nombre_archivo_nuevo = str(df_nro_cliente)+"_"+str(bill_number)+"_"+año+".pdf"
                    
                shutil.move(f'{downloads_folder}\\{nombre_archivo}',f'{folder_path}\\{nombre_archivo_nuevo}')
                descargado = False
              
                print("No se encontró el elemento con el id especificado...")
                print('----------------------------------------------------------------------')

                print('Conteo de documentos: ', i)
                print('----------------------------------------------------------------------')                

                print('----------------------------------------------------------------------')
                        
                time.sleep(10)
                print('pasamos al siguiente archivo')
            i +=1

    
    def buscar_texto_indirecto_y_moverse(self,check_texto,lista,sum_pos):
        try:
            posicion = lista.index(check_texto)
            variable = lista[posicion+sum_pos]
        except:
            print('elemento no se encuentra disponible')
            variable = ''
            
            
        return variable

    def buscar_texto_indirecto_y_moverse_evaluar(self,check_texto,lista):
        try:
            posicion_12 = lista.index(check_texto)
            proximo_estimada = lista[posicion_12+4]
            if proximo_estimada == 'proporcionalidad de precios del inicio de período':
                variable = lista[posicion_12+7]
            elif proximo_estimada != 'proporcionalidad de precios del inicio de período':
                variable = lista[posicion_12+4] 
        except:
            print('elemento no se encuentra disponible')
            variable = '-'
            
        return variable

    def buscar_texto_indirecto_moverse_formato_fecha(self,check_texto,lista,sum_pos):
        try:
            posicion = lista.index(check_texto)
            variable_bruto = lista[posicion+sum_pos]
            variable = variable_bruto.replace("/", "-")
        except:
            print('elemento no se encuentra disponible')
            variable = ''
            
            
        return variable

    def buscar_texto_indirecto_moverse_formato_alphanumerico(self,check_texto,lista,sum_pos):
        try:
            posicion = lista.index(check_texto)
            variable_bruto = lista[posicion+sum_pos]
            variable = variable_bruto.replace(".", "")
        except:
            print('elemento no se encuentra disponible')
            variable = ''
            
            
        return variable

    def buscar_texto_indirecto_moverse_segunda_validacion(self,check_texto,lista,sum_pos):
 
        try:
            posicion_10 = lista.index(check_texto)
            fecha_actual_bruto = lista[posicion_10+sum_pos]
            texto_sin_espacios = fecha_actual_bruto.replace(" ", "")
                        
            texto_a_verificar = 'Estimadoclientesuconsumoseencuentra'
            for idx, elemento in enumerate(lista):
                if texto_a_verificar in texto_sin_espacios:
                    fecha_actual_bruta = lista[posicion_10+6]
                    variable = fecha_actual_bruta.replace("/", "-")
                    break
                else:
                    fecha_actual_bruta = lista[posicion_10+2]
                    variable = fecha_actual_bruta.replace("/", "-")
                    break
                
        except:
            print('elemento no se encuentra disponible')
            lista = '-'
            
            
        return variable
    
    def archivos(self):
        
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia = [elemento.strip() for elemento in texto_completo.split('\n')]

                    #Posicion 0
                    
                    factura_elec = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,-3)
                        
                    #Posicion 1
                    razon_social = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,0)
                    
                    #Posicion 2
                    direccion = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,-1)
                    
                    #Posicion 3
                    comuna = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,2)

                    #Posicion 4
                    giro = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,1)
                    
                    #Posicion 5
                    rut = self.buscar_texto_indirecto_y_moverse('ABASTIBLE S.A.',lista_limpia,-2)
                    
                    #Posicion 6
                    n_cliente = self.buscar_texto_indirecto_y_moverse('Comuna',lista_limpia,1)
                        
                    #Posicion 7
                    fecha_emision = self.buscar_texto_indirecto_moverse_formato_fecha('Comuna:',lista_limpia,2)

                    #Posicion 8
                    ruta = self.buscar_texto_indirecto_y_moverse('Comuna:',lista_limpia,3)

                    #Posicion 9
                    fecha_vencimiento_bruto = lista_limpia[4]
                    fecha_vencimiento = fecha_vencimiento_bruto.replace("/", "-")
                    
                    #Posicion 10
                    fecha_actual = self.buscar_texto_indirecto_moverse_segunda_validacion('Acogido a Convenio  PAC Banco BANCO DE',lista_limpia,2)

                    #Posicion 11
                    fecha_anterior = self.buscar_texto_indirecto_moverse_formato_fecha('Acogido a Convenio  PAC Banco BANCO DE',lista_limpia,3)

                    #Posicion 12
                    proximo_estimada_bruto = self.buscar_texto_indirecto_y_moverse_evaluar('Acogido a Convenio  PAC Banco BANCO DE',lista_limpia)
                    proximo_estimada = proximo_estimada_bruto.replace('/','-')
                    
                    #Posicion 13
                    lectura_actual = lista_limpia[14]

                    #Posicion 14
                    lectura_anterior = lista_limpia[15]
                    
                    #Posicion 15
                    a_facturar = lista_limpia[17]
                    
                    #Posicion 16
                    clave_lectura = self.buscar_texto_indirecto_y_moverse('Clave de Lectura',lista_limpia,1)

                    #Posicion 17
                    n_medidor = self.buscar_texto_indirecto_y_moverse('Número Medidor',lista_limpia,1)

                    #Posicion 18
                    diametro_bruto = self.buscar_texto_indirecto_y_moverse('Número Medidor',lista_limpia,1)
                    diametro = diametro_bruto.rstrip(" mm")

                    #Posicion 19
                    factor_cobro = self.buscar_texto_indirecto_y_moverse('Factor de Cobro',lista_limpia,1)

                    #Posicion 20
                    limite_sobreconsumo_bruto = self.buscar_texto_indirecto_y_moverse('Límite de Sobreconsumo',lista_limpia,1)
                    limite_sobreconsumo = limite_sobreconsumo_bruto.rstrip("m3")

                    #Posicion 21
                    cargo_fijo = self.buscar_texto_indirecto_y_moverse('Cargo Fijo',lista_limpia,1)

                    #Posicion 22
                    consumo_agua = self.buscar_texto_indirecto_moverse_formato_alphanumerico('Consumo Agua',lista_limpia,1)

                    #Posicion 23
                    sencillo_anterior = self.buscar_texto_indirecto_y_moverse('Sencillo Anterior',lista_limpia,1)

                    #Posicion 24
                    monto_neto = self.buscar_texto_indirecto_moverse_formato_alphanumerico('Monto Neto',lista_limpia,1)

                    #Posicion 25
                    iva = self.buscar_texto_indirecto_moverse_formato_alphanumerico('19% IVA',lista_limpia,-1)

                    #Posicion 26
                    total_a_pagar_bruto = lista_limpia[9]
                    total_a_pagar = total_a_pagar_bruto.replace(" ", "").replace("$", "").replace(".", "")

                    #Posicion extra
                    sencillo_actual = self.buscar_texto_indirecto_y_moverse('Sencillo Actual',lista_limpia,1)

                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_agua = libro['Agua']
                    
                ultima_fila = hoja_agua.max_row
                
                #Los datos mas importantes
                hoja_agua.cell(row=ultima_fila+1,column=1).value = 1
                hoja_agua.cell(row=ultima_fila+1,column=2).value = int(factura_elec)
                
                #Primera tabla traspasada a excel
                #hoja_agua.cell(row=ultima_fila+1,column=3).value = razon_social
                hoja_agua.cell(row=ultima_fila+1,column=6).value = direccion
                hoja_agua.cell(row=ultima_fila+1,column=80).value = comuna
                hoja_agua.cell(row=ultima_fila+1,column=7).value = giro
                hoja_agua.cell(row=ultima_fila+1,column=8).value = rut
                hoja_agua.cell(row=ultima_fila+1,column=9).value = n_cliente
                hoja_agua.cell(row=ultima_fila+1,column=10).value = fecha_emision


                hoja_agua.cell(row=ultima_fila+1,column=64).value = ruta
                hoja_agua.cell(row=ultima_fila+1,column=11).value = fecha_vencimiento
                hoja_agua.cell(row=ultima_fila+1,column=57).value = fecha_actual
                hoja_agua.cell(row=ultima_fila+1,column=58).value = fecha_anterior
                hoja_agua.cell(row=ultima_fila+1,column=54).value = proximo_estimada
                hoja_agua.cell(row=ultima_fila+1,column=56).value = int(lectura_actual)
                hoja_agua.cell(row=ultima_fila+1,column=58).value = int(lectura_anterior)

                hoja_agua.cell(row=ultima_fila+1,column=60).value = (a_facturar)
                hoja_agua.cell(row=ultima_fila+1,column=59).value = clave_lectura
                hoja_agua.cell(row=ultima_fila+1,column=52).value = int(n_medidor)
                hoja_agua.cell(row=ultima_fila+1,column=53).value = int(diametro)
                hoja_agua.cell(row=ultima_fila+1,column=61).value = factor_cobro

                if limite_sobreconsumo != '':
                    hoja_agua.cell(row=ultima_fila+1,column=63).value = int(limite_sobreconsumo)
                elif limite_sobreconsumo == '':
                    hoja_agua.cell(row=ultima_fila+1,column=63).value = 0
                    
                #hoja_agua.cell(row=ultima_fila+1,column=8).value = cargo_fijo
                hoja_agua.cell(row=ultima_fila+1,column=15).value = int(consumo_agua)
                hoja_agua.cell(row=ultima_fila+1,column=82).value = sencillo_anterior
                if sencillo_anterior != '':
                    hoja_agua.cell(row=ultima_fila+1,column=82).value = int(sencillo_anterior)
                elif sencillo_anterior == '':
                    hoja_agua.cell(row=ultima_fila+1,column=82).value = ''

                hoja_agua.cell(row=ultima_fila+1,column=42).value = int(monto_neto)
                hoja_agua.cell(row=ultima_fila+1,column=43).value = int(iva)
                hoja_agua.cell(row=ultima_fila+1,column=46).value = int(total_a_pagar)
                if sencillo_actual != '':
                    hoja_agua.cell(row=ultima_fila+1,column=83).value = int(sencillo_actual)
                elif sencillo_actual == '':
                    hoja_agua.cell(row=ultima_fila+1,column=83).value = ''
                
            
                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
                #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
                shutil.copy(archivo, output_path+n_cliente+factura_elec+'.pdf')
                print('-----')

        #Obtenemos los archivos de la carpeta input
        archivos_en_carpeta = os.listdir(folder_path)

        # Iterar sobre los archivos y eliminarlos
        for archivo in archivos_en_carpeta:
            ruta_archivo = os.path.join(folder_path, archivo)
            if os.path.isfile(ruta_archivo):
                os.remove(ruta_archivo)
                    

                    
                    
    