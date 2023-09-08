from codigo.app_esval import Scraper_Esval
#import smtplib
from openpyxl import load_workbook

def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':
    
    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = 'C:\\roda\\esval-portal\\config\\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[0].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'
    
    scraper = Scraper_Esval(url, email, password, driver_path)
    #Primer ingreso, año actual
    print('Hacemos scrapping al portal...')
    print('----------------------------------------------------------------------')

    scraper.login()
    print('Hacemos login en el portal...')
    print('----------------------------------------------------------------------')
    
    scraper.scrapping_esval()
    print('Hacemos scrapping al portal...')
    print('----------------------------------------------------------------------')
            
    scraper.archivos()
    print('Procesando datos hacia Planilla Formato...')
    print('----------------------------------------------------------------------')
    
    scraper.close()
    print('Cerramos el bot...')
    print('----------------------------------------------------------------------')